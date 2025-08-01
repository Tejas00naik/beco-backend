"""Service for generating and uploading SAP-compatible XLSX exports."""

import logging
import pandas as pd
import os
import tempfile
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from uuid import uuid4

from src.config import DEFAULT_GCS_BUCKET_NAME
from src.repositories.firestore_dao import FirestoreDAO
from src.repositories.payment_advice_repository import PaymentAdviceRepository
from src.external_apis.gcp.gcs_uploader import GCSUploader
from src.models.schemas import PaymentAdvice, PaymentAdviceLine

logger = logging.getLogger(__name__)

class SAPExportService:
    """Service for generating and uploading SAP-compatible XLSX exports from payment advice lines."""
    
    def __init__(self, dao: FirestoreDAO):
        """Initialize with Firestore DAO."""
        self.dao = dao
        self.payment_advice_repo = PaymentAdviceRepository(dao)
        # Use the configured GCS bucket from environment or DEFAULT_GCS_BUCKET_NAME from config
        bucket_name = os.environ.get("GCP_STORAGE_BUCKET", DEFAULT_GCS_BUCKET_NAME)
        self.gcs_uploader = GCSUploader(bucket_name)
        
    # BP code lookup is now handled separately by the AccountEnrichmentService
            
    async def get_payment_advice_lines(self, payment_advice_uuid: str) -> List[Dict[str, Any]]:
        """
        Get all payment advice lines for a payment advice.
        
        Args:
            payment_advice_uuid: UUID of the payment advice
            
        Returns:
            List of payment advice line objects
        """
        try:
            # Query payment advice lines with the given payment advice UUID
            lines = await self.dao.query_documents(
                "paymentadvice_lines", 
                [("payment_advice_uuid", "==", payment_advice_uuid)]
            )
            
            logger.info(f"Found {len(lines)} payment advice lines for {payment_advice_uuid}")
            return lines
        except Exception as e:
            logger.error(f"Error getting payment advice lines for {payment_advice_uuid}: {str(e)}")
            return []
            
    async def get_payment_advice(self, payment_advice_uuid: str) -> Optional[Dict[str, Any]]:
        """
        Get a payment advice by UUID.
        
        Args:
            payment_advice_uuid: UUID of the payment advice
            
        Returns:
            Payment advice object if found, None otherwise
        """
        try:
            # Get payment advice
            payment_advice = await self.dao.get_document("payment_advice", payment_advice_uuid)
            if not payment_advice:
                logger.error(f"Payment advice {payment_advice_uuid} not found")
                return None
                
            return payment_advice
        except Exception as e:
            logger.error(f"Error getting payment advice {payment_advice_uuid}: {str(e)}")
            return None
            
    def map_payment_advice_lines_to_sap_format(
        self, 
        lines: List[Dict[str, Any]], 
        payment_advice: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """
        Map payment advice lines to SAP format for SAP B1 Excel import.
        
        Args:
            lines: List of payment advice lines
            payment_advice: Payment advice object
            
        Returns:
            List of mapped SAP rows
        """
        sap_rows = []
        payment_date = payment_advice.get("payment_advice_date")
        logger.info(f"Original payment_advice_date from Firestore: {payment_date}, type: {type(payment_date)}")
        
        if isinstance(payment_date, str):
            try:
                # Try multiple date formats since data in Firestore might be in different formats
                # First try YYYY-MM-DD format
                try:
                    payment_date = datetime.strptime(payment_date, "%Y-%m-%d").date()
                    logger.info(f"Converted payment_date from YYYY-MM-DD format: {payment_date}")
                except ValueError:
                    # Try DD/MM/YYYY format
                    try:
                        payment_date = datetime.strptime(payment_date, "%d/%m/%Y").date()
                        logger.info(f"Converted payment_date from DD/MM/YYYY format: {payment_date}")
                    except ValueError:
                        # Try MM/DD/YYYY format as last resort
                        payment_date = datetime.strptime(payment_date, "%m/%d/%Y").date()
                        logger.info(f"Converted payment_date from MM/DD/YYYY format: {payment_date}")
            except Exception as e:
                logger.warning(f"Failed to parse payment date: {payment_date}, {str(e)}")
                payment_date = None
        elif isinstance(payment_date, datetime):
            # If it's already a datetime object, just get the date part
            payment_date = payment_date.date()
            logger.info(f"Using datetime object directly: {payment_date}")
                
        # Map payment date to SAP format YYYY-MM-DD as requested
        payment_date_str = payment_date.strftime("%Y-%m-%d") if payment_date else ""
        logger.info(f"Final payment_date_str for SAP format: '{payment_date_str}'")
        
        # If payment_date_str is empty, try a fallback approach
        if not payment_date_str:
            logger.warning("Payment date string is empty, using fallback to current date")
            payment_date_str = datetime.now().strftime("%Y-%m-%d")
        
        # Add header row - minimal header as per example
        header_row = {
            "Record Type": "H",
            "Series": "JE25/",
            "Posting Date": payment_date_str,
            "Due Date": payment_date_str,
            "Document Date": payment_date_str,
            "Remarks": "",
            "Indicator Code": "",
            "Project Code": "",
            "Transaction Code": "",
            "Reference 1": "",
            "Reference 2": "",
            "Reference 3": "",
            "Generate Reg. No. or Not": "",
            "Material Type": "",
            "Loc.": "",
            "Account Type": "",
            "Account Code/BP Code": "",
            "Debit Amount": "",
            "Credit Amount": "",
            "Branch Name": "",
            "Reference 1 (in Lines)": "",
            "Reference 2 (in Lines)": "",
            "Reference 3 (in Lines)": ""
        }
        sap_rows.append(header_row)
        
        # Add line rows
        for line in lines:
            # Determine debit/credit amounts
            dr_amt = ""
            cr_amt = ""
            
            amount = line.get("amount", 0)
            dr_cr = line.get("dr_cr", "")
            
            if dr_cr.lower() == "dr":
                dr_amt = str(abs(float(amount))) if amount else ""
            elif dr_cr.lower() == "cr":
                cr_amt = str(abs(float(amount))) if amount else ""
            else:
                # If dr_cr not specified, use sign of amount
                if amount:
                    if float(amount) > 0:
                        dr_amt = str(abs(float(amount)))
                    else:
                        cr_amt = str(abs(float(amount)))
            
            # Map payment advice line to SAP format
            sap_row = {
                "Record Type": "L",
                "Series": "",  # Empty for lines as per example
                "Posting Date": payment_date_str,
                "Due Date": payment_date_str,
                "Document Date": payment_date_str,
                "Remarks": "",
                "Indicator Code": "",
                "Project Code": "",
                "Transaction Code": "",
                "Reference 1": "",
                "Reference 2": "",
                "Reference 3": "",
                "Generate Reg. No. or Not": "",
                "Material Type": "",
                "Loc.": "",
                "Account Type": line.get("account_type", "").upper(),  # GL or BP
                "Account Code/BP Code": line.get("bp_code") or line.get("gl_code", ""),
                "Debit Amount": dr_amt,
                "Credit Amount": cr_amt,
                "Branch Name": line.get("branch_name", "MAHARASHTRA").upper(),  # State name in uppercase
                "Reference 1 (in Lines)": line.get("ref_1", ""),
                "Reference 2 (in Lines)": line.get("ref_2", ""),
                "Reference 3 (in Lines)": line.get("ref_3", "")
            }
            sap_rows.append(sap_row)
            
        return sap_rows
            
    async def generate_sap_excel(
        self, 
        payment_advice_uuid: str
    ) -> Tuple[Optional[str], Optional[str]]:
        """
        Generate SAP Excel file for a payment advice.
        
        Args:
            payment_advice_uuid: UUID of the payment advice
            
        Returns:
            Tuple of (local file path, filename) if successful, (None, None) otherwise
        """
        try:
            # Get payment advice
            payment_advice = await self.get_payment_advice(payment_advice_uuid)
            if not payment_advice:
                return None, None
                
            # Get payment advice lines
            lines = await self.get_payment_advice_lines(payment_advice_uuid)
            if not lines:
                logger.warning(f"No payment advice lines found for {payment_advice_uuid}")
                return None, None
                
            # Map payment advice lines to SAP format - BP/GL codes should already be enriched
            sap_rows = self.map_payment_advice_lines_to_sap_format(lines, payment_advice)
            
            # Create DataFrame
            df = pd.DataFrame(sap_rows)
            
            # Generate temporary file
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                temp_file_path = temp_file.name
                
            # Write DataFrame to Excel without index and without header row
            df.to_excel(temp_file_path, index=False, header=False, engine="openpyxl")
            
            # Optional: Format numbers to avoid scientific notation
            try:
                from openpyxl import load_workbook
                wb = load_workbook(temp_file_path)
                ws = wb.active
                
                # Format number columns properly
                for row in ws.iter_rows(min_row=2):
                    # Debit Amount column
                    if row[17].value and row[17].value != "":
                        row[17].number_format = '#,##0.00'
                    
                    # Credit Amount column
                    if row[18].value and row[18].value != "":
                        row[18].number_format = '#,##0.00'
                        
                wb.save(temp_file_path)
            except Exception as e:
                logger.warning(f"Error formatting Excel: {str(e)}")
                # Continue without formatting - the basic export will still work
            
            # Generate filename
            payment_advice_number = payment_advice.get("payment_advice_number", "unknown")
            payment_date = payment_advice.get("payment_advice_date")
            if isinstance(payment_date, datetime):
                date_str = payment_date.strftime("%Y%m%d")
            else:
                date_str = datetime.now().strftime("%Y%m%d")
                
            filename = f"SAP_Export_{payment_advice_number}_{date_str}.xlsx"
            
            logger.info(f"Generated SAP Excel file at {temp_file_path} with filename {filename}")
            return temp_file_path, filename
        except Exception as e:
            logger.error(f"Error generating SAP Excel for {payment_advice_uuid}: {str(e)}")
            return None, None
            
    async def upload_to_gcp(self, file_path: str, filename: str) -> Optional[str]:
        """
        Upload a file to GCP Storage and generate a presigned URL.
        
        Args:
            file_path: Local path to file
            filename: Name for the uploaded file
            
        Returns:
            Presigned URL if successful, None otherwise
        """
        try:
            # Generate a unique folder name (to avoid conflicts)
            destination_folder = f"sap_exports/{str(uuid4())}"
            
            # Upload file to GCP and get presigned URL (valid for 7 days)
            url = self.gcs_uploader.upload_and_get_signed_url(
                file_path=file_path,
                destination_folder=destination_folder,
                filename=filename,
                expiration_days=7
            )
            
            if not url:
                logger.error(f"Failed to upload {file_path} to GCP")
                return None
                
            logger.info(f"Uploaded {file_path} to GCP and generated presigned URL valid for 7 days")
            return url
        except Exception as e:
            logger.error(f"Error uploading {file_path} to GCP: {str(e)}")
            return None
        finally:
            # Clean up temporary file
            try:
                if os.path.exists(file_path):
                    os.unlink(file_path)
            except Exception as e:
                logger.error(f"Error cleaning up {file_path}: {str(e)}")
            
    async def update_payment_advice_with_url(
        self, 
        payment_advice_uuid: str, 
        url: str
    ) -> bool:
        """
        Update payment advice with SAP export URL.
        
        Args:
            payment_advice_uuid: UUID of the payment advice
            url: Presigned URL for the SAP export
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Update payment advice with URL
            updates = {
                "sap_export_url": url,
                "updated_at": datetime.utcnow()
            }
            
            # Update payment advice in Firestore
            await self.dao.update_document("payment_advice", payment_advice_uuid, updates)
            
            logger.info(f"Updated payment advice {payment_advice_uuid} with SAP export URL")
            return True
        except Exception as e:
            logger.error(f"Error updating payment advice {payment_advice_uuid} with URL: {str(e)}")
            return False
            
    async def process_payment_advice_export(self, payment_advice_uuid: str) -> Optional[str]:
        """
        Process payment advice export: generate SAP Excel, upload to GCP, update payment advice.
        
        Args:
            payment_advice_uuid: UUID of the payment advice
            
        Returns:
            Presigned URL if successful, None otherwise
        """
        try:
            # Generate SAP Excel
            file_path, filename = await self.generate_sap_excel(payment_advice_uuid)
            if not file_path or not filename:
                logger.error(f"Failed to generate SAP Excel for {payment_advice_uuid}")
                return None
                
            # Upload to GCP
            url = await self.upload_to_gcp(file_path, filename)
            if not url:
                logger.error(f"Failed to upload SAP Excel to GCP for {payment_advice_uuid}")
                return None
                
            # Update payment advice with URL
            success = await self.update_payment_advice_with_url(payment_advice_uuid, url)
            if not success:
                logger.error(f"Failed to update payment advice {payment_advice_uuid} with URL")
                return None
                
            logger.info(f"Successfully processed SAP export for {payment_advice_uuid}")
            return url
        except Exception as e:
            logger.error(f"Error processing SAP export for {payment_advice_uuid}: {str(e)}")
            return None
